"""Build a second test workbook, deliberately unlike the first.

Everything was tested against one file — GROUP 3's production plan, counting images — and
Decision 3's whole claim is that Lumina works on a workbook nobody has seen. This one
counts videos, has two sheets so the agent must ask which, four studios and twenty-eight
editors so there is more than one useful breakdown, and a March collapse so a shortfall
means something. It carries the same hazards as the real workbook: dashes where a figure
was never filled in, an unlabelled grand total, trailing padding, and rows with a date and
nothing else.

Kept as a script rather than a file so it is reviewable, reproducible and cannot go stale
against the code. Run it and upload the result:

    python lumina/backend/testdata/make_video_plan.py

It found five defects in the twenty minutes after it was first written, every one of which
had passed against the original file. Section 43 of the decision log records them.
"""

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

random.seed(7)  # so the same file comes out every time

STUDIOS = ["Manila", "Cebu", "Davao", "Iloilo"]
EDITORS = [f"Editor {i:02d}" for i in range(1, 29)]  # enough to need a top ten

# A strong start, a collapse in March, a partial recovery — so that a shortfall chart has
# something to say and a single overall figure hides it.
SHAPE = {1: 1.15, 2: 1.30, 3: 0.22, 4: 0.78}

HEADINGS = [
    "Work Date", "Week", "Studio", "Editor",
    "Planned Videos", "Completed Videos",
    "Achievement Rate", "Variance",
    "Running Planned", "Running Completed",
    "Notes",
]


def build() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Video Production Plan"
    ws.append(HEADINGS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    run_planned = run_completed = 0
    rows = 0
    day = date(2026, 1, 5)
    while day < date(2026, 5, 1):
        if day.weekday() >= 5:  # weekdays only, as a real plan would be
            day += timedelta(days=1)
            continue
        for studio in STUDIOS:
            planned = random.choice([6, 8, 10, 12])
            completed = max(
                0, round(planned * SHAPE[day.month] * random.uniform(0.8, 1.2))
            )
            run_planned += planned
            run_completed += completed
            ws.append([
                day,
                f"W{day.isocalendar().week:02d}",
                studio,
                random.choice(EDITORS),
                # Some planned figures were never filled in, exactly as the real
                # workbook leaves them.
                "-" if rows % 37 == 0 else planned,
                completed,
                (completed / planned) if planned else None,
                completed - planned,
                run_planned,
                run_completed,
                "" if rows % 23 else "rework",
            ])
            rows += 1
        day += timedelta(days=1)

    # Two rows carrying a date and no figures, an unlabelled grand total, and padding.
    ws.append([date(2026, 5, 4), "W19", "Manila", "Editor 01", *[None] * 6, ""])
    ws.append([date(2026, 5, 5), "W19", "Cebu", "Editor 02", *[None] * 6, ""])
    ws.append([
        None, None, None, None, run_planned, run_completed, None, 0,
        run_planned, run_completed, "TOTAL",
    ])
    for _ in range(4):
        ws.append([None] * len(HEADINGS))

    # A second sheet, so the agent has to ask which one the dashboard is about.
    log = wb.create_sheet("Upload Log")
    log.append(["Upload ID", "Work Date", "Studio", "File Name", "Seconds"])
    for i in range(1, 601):
        log.append([
            f"UP{i:05d}",
            date(2026, 1, 5) + timedelta(days=random.randint(0, 110)),
            random.choice(STUDIOS),
            f"clip_{i:05d}.mp4",
            random.randint(20, 900),
        ])

    return wb


if __name__ == "__main__":
    out = Path.home() / "Downloads" / "Video Production Plan - TEST.xlsx"
    build().save(out)
    print(f"written: {out}")
