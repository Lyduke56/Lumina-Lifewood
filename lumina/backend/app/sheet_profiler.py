"""Describe what is actually in a spreadsheet, so the agent can reason about it.

This is the first of the six tools in Decision 8, and everything else depends on it.
The agent decides what a column *means* (Decision 3) and which breakdowns are worth
offering (Decision 6) from the facts gathered here, and the chart-choosing step reads
it so that it stops proposing charts for figures a workbook does not contain.

Nothing here interprets meaning. It reports observations and flags hazards; judgement
belongs to the agent, and confirmation to the customer.
"""

from __future__ import annotations

import datetime as dt
from collections import Counter, deque
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# Rows kept at the start to locate the headings, and at the end to spot padding.
HEAD_ROWS = 40
TAIL_ROWS = 60

# Distinct values are counted across the whole sheet, but counting stops here. The only
# question being answered is "few enough to chart?", and once a column is past the cap
# the exact figure changes no decision. Counting a sample instead would be wrong: real
# workbooks are sorted, so the opening rows badly understate how much variety follows —
# the official workbook shows 16 participants in its first 2,000 rows and many more
# after that.
DISTINCT_CAP = 200

# More distinct values than this makes an unreadable chart, so the column is offered as
# a "top N" breakdown rather than a plain one (Decision 6).
MAX_GROUPS_FOR_BREAKDOWN = 25

# Above this, a column is almost certainly an identifier rather than a label, and
# grouping by it would produce roughly one bucket per row.
IDENTIFIER_UNIQUENESS = 0.9


@dataclass
class ColumnProfile:
    """What was observed about one column. No interpretation of meaning."""

    position: int  # 1-based, as a person would count columns in Excel
    heading: str | None
    kinds: dict[str, int]  # "number" / "date" / "text" / "blank" -> count
    distinct: int
    distinct_capped: bool  # True when there are at least `distinct` — possibly many more
    examples: list
    blank_fraction: float
    non_numeric_examples: list  # e.g. the "-" placeholders in the official workbook

    @property
    def mostly(self) -> str:
        real = {k: v for k, v in self.kinds.items() if k != "blank"}
        return max(real, key=real.get) if real else "blank"

    @property
    def populated(self) -> int:
        return sum(v for k, v in self.kinds.items() if k != "blank")

    @property
    def is_identifier(self) -> bool:
        if self.distinct_capped:
            return True
        return self.populated > 0 and self.distinct / self.populated >= IDENTIFIER_UNIQUENESS

    @property
    def breakdown_suitability(self) -> str:
        """Whether grouping by this column would produce a readable chart."""
        if self.mostly in ("number", "date", "blank") or self.is_identifier:
            return "unsuitable"
        if self.distinct < 2:
            return "unsuitable"
        return "good" if self.distinct <= MAX_GROUPS_FOR_BREAKDOWN else "top-n-only"


@dataclass
class SheetProfile:
    name: str
    total_rows: int
    header_row: int | None
    data_starts_at: int | None
    data_row_count: int
    blank_rows_at_end: int = 0
    # Set when a grand total is found sitting among the ordinary rows. Reported as a
    # row number rather than only as prose, so that summarising can actually skip it.
    suspected_total_row: int | None = None
    columns: list[ColumnProfile] = field(default_factory=list)
    empty_columns: list[int] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _kind(value) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return "blank"
    if isinstance(value, (dt.datetime, dt.date)):
        return "date"
    if isinstance(value, bool):
        return "text"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


def list_sheets(path: str | Path) -> list[dict]:
    """Every sheet with its size — so the agent can ask which ones are wanted.

    Replaces the previous behaviour of rejecting any workbook with more than one
    sheet, which turned the official Lifewood workbook away at the door.
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return [
            {
                "name": name,
                "rows": wb[name].max_row or 0,
                "columns": wb[name].max_column or 0,
            }
            for name in wb.sheetnames
            if not name.startswith("WpsReserved_")
        ]
    finally:
        wb.close()


def _find_header_and_data(rows: list[tuple]) -> tuple[int | None, int | None]:
    """Locate the heading row and the first row of real data.

    The first row holding several dates or numbers is taken as data; the last
    predominantly-text row above it is taken as the headings.
    """
    for i, row in enumerate(rows):
        kinds = Counter(_kind(v) for v in row)
        if kinds["date"] + kinds["number"] >= 2:
            header = None
            for j in range(i - 1, -1, -1):
                above = Counter(_kind(v) for v in rows[j])
                if above["text"] >= max(1, above["number"] + above["date"]):
                    header = j + 1
                    break
            return header, i + 1
    return None, None


class _ColumnAccumulator:
    """Running tallies for one column, so the sheet is read exactly once."""

    __slots__ = ("kinds", "seen", "capped", "examples", "odd_text", "first", "last")

    def __init__(self) -> None:
        self.kinds: Counter = Counter()
        self.seen: set[str] = set()
        self.capped = False
        self.examples: list = []
        self.odd_text: set[str] = set()
        self.first = None
        self.last = None

    def add(self, value) -> None:
        kind = _kind(value)
        self.kinds[kind] += 1
        if kind == "blank":
            return
        if not self.capped:
            if len(self.seen) < DISTINCT_CAP:
                self.seen.add(str(value))
            else:
                self.capped = True  # past this the exact count changes no decision
        if len(self.examples) < 3 and value not in self.examples:
            self.examples.append(value)
        if kind == "text" and len(str(value).strip()) <= 3 and len(self.odd_text) < 3:
            self.odd_text.add(str(value))
        if kind == "number":
            if self.first is None:
                self.first = value
            self.last = value


def profile_sheet(path: str | Path, sheet_name: str) -> SheetProfile:
    """Describe one sheet: its columns, what they hold, and anything hazardous.

    Reads every row once. Distinct values are counted across the whole sheet rather
    than sampled, because sampling a sorted workbook understates its variety and would
    let an unusable column be offered as a breakdown.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        total_rows = ws.max_row or 0

        tail: deque = deque(maxlen=TAIL_ROWS)
        columns: dict[int, _ColumnAccumulator] = {}
        data_rows = 0
        first_data_row: tuple | None = None

        trailing_blank = 0

        def take(row: tuple) -> None:
            """Fold one data row into the running tallies."""
            nonlocal data_rows, trailing_blank
            data_rows += 1
            # Spreadsheets are routinely padded with empty rows below the real data —
            # the official workbook has 22. Counting them would overstate how much
            # data there is, so the run at the foot is tracked and removed afterwards.
            trailing_blank = (
                trailing_blank + 1 if all(_kind(v) == "blank" for v in row) else 0
            )
            tail.append(row)
            for i, value in enumerate(row):
                columns.setdefault(i, _ColumnAccumulator()).add(value)

        stream = ws.iter_rows(values_only=True)

        # Buffer the opening rows first — where the data begins cannot be known until
        # they have been seen, and rows must not be consumed before that is settled.
        head: list[tuple] = []
        for row in stream:
            head.append(row)
            if len(head) >= HEAD_ROWS:
                break
        header_row, data_start = _find_header_and_data(head)

        if data_start is not None:
            for n, row in enumerate(head, start=1):
                if n >= data_start:
                    if first_data_row is None:
                        first_data_row = row
                    take(row)
            for row in stream:
                take(row)
    finally:
        wb.close()

    profile = SheetProfile(
        name=sheet_name,
        total_rows=total_rows,
        header_row=header_row,
        data_starts_at=data_start,
        data_row_count=data_rows - trailing_blank,
        blank_rows_at_end=trailing_blank,
    )
    if not head:
        profile.warnings.append("The sheet is empty.")
        return profile
    if data_start is None:
        profile.warnings.append(
            "No row of figures could be found, so this may not be a data sheet."
        )
        return profile

    headings = head[header_row - 1] if header_row and header_row <= len(head) else ()
    for i in sorted(columns):
        acc = columns[i]
        if acc.kinds.get("blank", 0) == sum(acc.kinds.values()):
            profile.empty_columns.append(i + 1)
            continue
        heading = headings[i] if i < len(headings) else None
        total = sum(acc.kinds.values())
        profile.columns.append(
            ColumnProfile(
                position=i + 1,
                heading=str(heading).strip() if heading is not None else None,
                kinds=dict(acc.kinds),
                distinct=len(acc.seen),
                distinct_capped=acc.capped,
                examples=acc.examples,
                blank_fraction=acc.kinds.get("blank", 0) / total if total else 1.0,
                non_numeric_examples=sorted(acc.odd_text),
            )
        )

    profile.warnings.extend(_hazards(profile, first_data_row, tail, data_start))
    return profile


def _counter_columns(profile: SheetProfile) -> set[int]:
    """Columns that just number the rows — 1, 2, 3… — and carry no meaning.

    They have to be recognised or they mask genuinely empty rows: a padding row at the
    foot of a sheet still carries its row number, so it does not look empty.
    """
    return {
        c.position
        for c in profile.columns
        if c.mostly == "number"
        and (c.distinct_capped or c.distinct >= c.populated * IDENTIFIER_UNIQUENESS)
    }


def _hazards(
    profile: SheetProfile,
    first_data_row: tuple | None,
    tail: deque,
    data_start: int,
) -> list[str]:
    """Flag the things that quietly corrupt a dashboard if nobody notices them.

    Every one of these was found in the official Lifewood workbook.
    """
    warnings: list[str] = []
    ignorable = _counter_columns(profile)

    # A grand total sitting among the ordinary rows is read as one enormous day.
    numeric = [
        c
        for c in profile.columns
        if c.mostly == "number" and c.position not in ignorable
    ]
    if numeric and first_data_row is not None and tail:
        for c in numeric:
            i = c.position - 1
            top = first_data_row[i] if i < len(first_data_row) else None
            others = [
                r[i]
                for r in tail
                if r is not first_data_row
                and i < len(r)
                and isinstance(r[i], (int, float))
            ]
            if isinstance(top, (int, float)) and others and top > 5 * max(others):
                profile.suspected_total_row = data_start
                warnings.append(
                    f"Row {data_start} looks like a grand total rather than an ordinary "
                    f'row — under "{c.heading or f"column {c.position}"}" it is far '
                    f"larger than the rows below. Including it would distort every chart."
                )
                break

    if profile.blank_rows_at_end:
        warnings.append(
            f"The last {profile.blank_rows_at_end} row(s) are completely empty padding "
            f"and have been left out of the count above."
        )

    # A subtler variant: a row carrying a date, and perhaps a label such as the month,
    # but no actual figures. Row numbers and other counters are ignored, or such a row
    # would never look empty.
    date_col = next((c for c in profile.columns if c.mostly == "date"), None)
    if date_col:
        i = date_col.position - 1
        trailing = 0
        for row in list(reversed(tail))[profile.blank_rows_at_end :]:
            has_date = i < len(row) and _kind(row[i]) == "date"
            figures = [
                v
                for j, v in enumerate(row)
                if j != i and (j + 1) not in ignorable and _kind(v) == "number"
            ]
            if has_date and not figures:
                trailing += 1
            else:
                break
        if trailing:
            warnings.append(
                f"A further {trailing} row(s) carry a date but no figures, and should "
                f"probably be ignored too."
            )

    # Number columns holding "-" or similar cannot simply be added up.
    for c in profile.columns:
        if c.kinds.get("number") and c.non_numeric_examples:
            warnings.append(
                f'"{c.heading or f"column {c.position}"}" is mostly numbers but also '
                f"contains {', '.join(repr(x) for x in c.non_numeric_examples)}, which "
                f"cannot be added up."
            )

    if not profile.header_row:
        warnings.append("No heading row was found, so the columns are unnamed.")

    if profile.empty_columns:
        warnings.append(
            f"{len(profile.empty_columns)} column(s) are entirely empty and were ignored."
        )

    return warnings
