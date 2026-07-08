import json
from pathlib import Path

import openpyxl

from llm_client import ask

CANONICAL_FIELDS = [
    "target_quantity",
    "actual_quantity",
    "target_hours",
    "actual_hours",
]

import hashlib

CACHE_PATH = Path(__file__).resolve().parent / "column_mapping_cache.json"


def _labels_signature(labels: list[str | None]) -> str:
    joined = "|".join(label or "" for label in labels)
    return hashlib.sha256(joined.encode("utf-8")).hexdigest()


def _load_cache() -> dict:
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text())
    return {}


def _get_column_map(
    labels: list[str | None], sample_rows: list[tuple]
) -> dict[int, str]:
    """Use a previously-verified mapping if this exact layout has been seen before;
    otherwise call the LLM and print the result for manual review/promotion into the cache.
    """
    cache = _load_cache()
    signature = _labels_signature(labels)
    if signature in cache:
        return {int(k): v for k, v in cache[signature]["column_map"].items()}

    mapping = _map_columns_to_canonical(labels, sample_rows)
    entry = {
        signature: {
            "column_map": {str(k): v for k, v in mapping.items()},
            "labels": labels,
        }
    }
    print(
        f"NEW FILE LAYOUT (signature {signature[:8]}) — unverified, review before trusting it.\n"
        f"If correct, paste this entry into column_mapping_cache.json to skip the LLM next time:\n"
        + json.dumps(entry, indent=2)
    )
    return mapping


def load_production_plan(path: str | Path) -> list[dict]:
    """Read the single sheet in the workbook and return one dict per daily record."""
    wb = openpyxl.load_workbook(path, data_only=True)
    real_sheets = [
        name for name in wb.sheetnames if not name.startswith("WpsReserved_")
    ]
    if len(real_sheets) != 1:
        raise ValueError(
            f"Expected a file with exactly one sheet (the Production Plan itself), "
            f"found {len(real_sheets)}: {real_sheets}. "
            f"Remove any extra sheets before uploading."
        )
    ws = wb[real_sheets[0]]

    data_start_row = _find_data_start_row(ws)
    max_col = _find_real_max_col(ws, data_start_row)
    labels = _build_header_labels(ws, data_start_row, max_col)
    sample_rows = [
        row
        for row in ws.iter_rows(
            min_row=data_start_row,
            max_row=data_start_row + 6,
            max_col=max_col,
            values_only=True,
        )
        if hasattr(row[0], "year")
    ][:3]
    column_map = _get_column_map(labels, sample_rows)

    records = []
    for row in ws.iter_rows(min_row=data_start_row, max_col=max_col, values_only=True):
        date = row[0]
        if not hasattr(date, "year"):
            continue

        totals = {field: 0 for field in CANONICAL_FIELDS}
        for col_index, field in column_map.items():
            value = row[col_index - 1]
            if isinstance(value, (int, float)):
                totals[field] += value

        target_quantity = totals["target_quantity"]
        completion_rate = (
            (totals["actual_quantity"] / target_quantity) if target_quantity else None
        )

        raw = {
            labels[i]: (row[i].isoformat() if hasattr(row[i], "isoformat") else row[i])
            for i in range(len(labels))
            if labels[i] is not None and row[i] is not None
        }

        records.append(
            {
                "date": date.isoformat(),
                "target_quantity": totals["target_quantity"],
                "actual_quantity": totals["actual_quantity"],
                "target_hours": totals["target_hours"],
                "actual_hours": totals["actual_hours"],
                "completion_rate": completion_rate,
                "raw": raw,
            }
        )

    return records


def build_dashboard_preview(records: list[dict], visuals: list[dict]) -> tuple[dict, dict]:
    """Lightweight summary + chart series for the in-app preview panel, ahead of the full PBIP download."""
    total_target = sum(r["target_quantity"] or 0 for r in records)
    total_actual = sum(r["actual_quantity"] or 0 for r in records)

    layout_json = {
        "sections": [
            {
                "number": 1,
                "title": "Production Overview",
                "summary": {
                    "total_target_quantity": total_target,
                    "total_actual_quantity": total_actual,
                },
            }
        ]
    }
    chart_preview_json = {
        "visuals": visuals,
        "records": [
            {
                "date": r["date"],
                "target_quantity": r["target_quantity"],
                "actual_quantity": r["actual_quantity"],
                "target_hours": r["target_hours"],
                "actual_hours": r["actual_hours"],
                "completion_rate": r["completion_rate"],
            }
            for r in records
        ],
    }
    return layout_json, chart_preview_json


def _find_data_start_row(ws, date_col=1, max_scan=20):
    for r in range(2, max_scan):
        val = ws.cell(row=r, column=date_col).value
        if hasattr(val, "year"):
            return r
    raise ValueError("No row with a real date found in the first column.")


def _find_real_max_col(ws, data_start_row):
    last = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=data_start_row, column=col).value is not None:
            last = col
    return last


def _build_header_labels(ws, data_start_row, max_col):
    merged_lookup = {}
    for mr in ws.merged_cells.ranges:
        top_left = ws.cell(row=mr.min_row, column=mr.min_col).value
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merged_lookup[(row, col)] = top_left

    def resolved_row(r):
        return [
            merged_lookup.get((r, c), ws.cell(row=r, column=c).value)
            for c in range(1, max_col + 1)
        ]

    header_rows = []
    for r in range(1, data_start_row):
        is_full_width_title = any(
            mr.min_row <= r <= mr.max_row and mr.min_col == 1 and mr.max_col >= max_col
            for mr in ws.merged_cells.ranges
        )
        vals = resolved_row(r)
        non_none = [v for v in vals if v is not None]
        numeric_frac = (
            sum(
                1
                for v in non_none
                if isinstance(v, (int, float)) and not isinstance(v, bool)
            )
            / len(non_none)
            if non_none
            else 0
        )
        if not is_full_width_title and numeric_frac <= 0.5:
            header_rows.append(r)

    labels = []
    for col in range(1, max_col + 1):
        parts = []
        for row in header_rows:
            val = merged_lookup.get((row, col), ws.cell(row=row, column=col).value)
            if val is not None:
                text = str(val).replace("\n", " ").strip()
                if text and (not parts or parts[-1] != text):
                    parts.append(text)
        labels.append(" | ".join(parts) if parts else None)
    return labels


def _map_columns_to_canonical(
    labels: list[str | None], sample_rows: list[tuple]
) -> dict[int, str]:
    """Ask the LLM which columns feed which canonical field. Returns {column_index: field_name}."""
    numbered = "\n".join(
        f"{i}: {label}" for i, label in enumerate(labels, start=1) if label
    )
    sample_lines = "\n".join(
        f"Row {n}: "
        + ", ".join(
            f"col{i}={row[i - 1]}" for i in range(1, len(labels) + 1) if labels[i - 1]
        )
        for n, row in enumerate(sample_rows, start=1)
    )

    prompt = f"""You are mapping Excel column headers from a production-plan spreadsheet to a fixed canonical schema.

Column headers (numbered):
{numbered}

A few real data rows, to help disambiguate columns whose header text alone is ambiguous
(target/planned columns tend to stay constant or change rarely across rows; actual columns
vary day to day):
{sample_lines}

Canonical fields:
- target_quantity: planned/target number of work units (e.g. ops, tasks) for the day
- actual_quantity: actual number of work units completed for the day
- target_hours: planned/target hours of work for the day
- actual_hours: actual hours of work completed for the day

A column maps to a canonical field if it holds that quantity, possibly for one team/category among
several (in which case multiple columns will map to the same field, one per team/category, and
their values get SUMMED). Some files break the same work into multiple related counts at different
granularities (e.g. a count of "operations" and a related count of "tasks" derived from those
operations). If a single team/category has more than one column that could represent the same
canonical field, choose only ONE of them — never map two columns from the same team/category to
the same canonical field, since that double-counts the same work.
Do not map cumulative/accumulative totals, rates, percentages, per-unit breakdowns, or metadata
like Month.

Respond with ONLY a JSON object mapping column numbers (as strings) to one of:
"target_quantity", "actual_quantity", "target_hours", "actual_hours".
Omit columns that don't map to any of these. No other text."""

    response = ask(prompt, temperature=0, use_fallback=False)
    try:
        start, end = response.find("{"), response.rfind("}")
        mapping = json.loads(response[start : end + 1])
    except (json.JSONDecodeError, ValueError) as e:
        raise ValueError(
            "We couldn't automatically understand this file's columns. "
            "Please double-check the file format, or try again in a moment."
        ) from e
    return {int(k): v for k, v in mapping.items() if v in CANONICAL_FIELDS}


if __name__ == "__main__":
    import sys

    records = load_production_plan(sys.argv[1])
    print(f"Parsed {len(records)} daily records")
    print("First:", records[0])
    print("Last:", records[-1])
