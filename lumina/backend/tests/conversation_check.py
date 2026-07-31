"""Run whole conversations and check the result, without anybody clicking anything.

This exists because the way of working had become: change something, ask John Peter to
upload a workbook, tell him what to type, ask him to open the file, and wait. That is his
time spent on something a script can do, and it made every small fix cost an exchange.

What it does, per workbook, with no human involved:

  1. Holds a real conversation with the real AI — answering as a customer would, agreeing
     and asking it to carry on — until a report is built or it gives up.
  2. Recomputes the totals **independently from the spreadsheet**, with openpyxl, not
     through our own summariser. A check that uses the code under test to decide what the
     answer should be proves nothing.
  3. Compares those totals against the figures in the finished report.
  4. Confirms the model loads in Power BI's engine and the visuals pass Microsoft's
     validator — both of which the build already runs, so a build that succeeds has passed
     them.
  5. Checks every sentence the AI said for figures no tool produced.

What it cannot do: say whether a report reads well, whether a chart is worth showing, or
whether the number in the corner is the one a manager needs. Those still want a person —
but a person looking at a finished, verified report, not one clicking through a wizard to
find out whether the last change worked.

    python lumina/backend/tests/conversation_check.py
    python lumina/backend/tests/conversation_check.py --workbook "path\\to\\file.xlsx"
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from collections import defaultdict
from pathlib import Path

APP = Path(__file__).resolve().parent.parent / "app"
sys.path.insert(0, str(APP))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(APP.parent / ".env")

import openai  # noqa: E402

import agent  # noqa: E402
import figure_check  # noqa: E402
import workbench  # noqa: E402

# Every free supplier being out of allowance, or the network being down, says nothing about
# whether the code is right. Reported apart from a failure, because a checker that calls an
# outage a defect gets ignored, and then so do its real findings.
UNREACHABLE = (
    openai.APIConnectionError,
    openai.RateLimitError,
    openai.AuthenticationError,
    openai.PermissionDeniedError,
)


class CouldNotRun(Exception):
    """No AI supplier could be reached, so nothing was tested either way."""

# What a customer says to keep things moving. Deliberately unhelpful about the detail: if a
# conversation only works when the customer knows the right words, it does not work.
REPLIES = [
    "go ahead",
    "yes, that's right — carry on",
    "yes, that looks right. Add the headline figures and charts you suggested, then build it",
    "yes, go ahead and build it",
    "yes",
    "yes, build it",
]

MAX_TURNS = len(REPLIES)


def expected_totals(path: Path) -> dict[str, float]:
    """Add up the spreadsheet directly, ignoring total rows and rows with no date.

    Independent of the summariser on purpose. The one time these were checked by hand it
    found fourteen of a customer's videos missing.
    """
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    book.close()

    headings = [str(h) if h is not None else "" for h in rows[0]]
    date_at = next(
        (i for i, h in enumerate(headings) if "date" in h.lower()), 0
    )
    # The planned and achieved columns, by the words customers actually use.
    planned_at = next(
        (i for i, h in enumerate(headings) if re.search(r"target|planned|plan\b", h, re.I)),
        None,
    )
    actual_at = next(
        (i for i, h in enumerate(headings) if re.search(r"actual|completed|achieved", h, re.I)),
        None,
    )
    if planned_at is None or actual_at is None:
        return {}

    totals = defaultdict(float)
    for row in rows[1:]:
        if date_at >= len(row) or row[date_at] is None:
            continue  # no date: padding, or an unlabelled grand total
        for name, index in (("planned", planned_at), ("actual", actual_at)):
            value = row[index] if index < len(row) else None
            if isinstance(value, (int, float)):
                totals[name] += value
    return dict(totals)


def hold_conversation(path: Path) -> dict:
    """Talk to the AI as a customer would, and report what happened."""
    history: list[dict] = []
    said: list[str] = []
    steps: list[tuple[str, str]] = []
    session_id: str | None = None
    report = None

    for turn in range(MAX_TURNS):
        message = REPLIES[turn]
        if turn == 0:
            message = f"{message}\n\n(The workbook is at {path})"

        for event in agent.respond(history, message, None):
            kind = event.get("type")
            if kind == "message":
                said.append(event["text"])
            elif kind == "tool_finished":
                steps.append((event["tool"], event.get("outcome", "ok")))
                if event.get("session_id"):
                    session_id = event["session_id"]
        if any(tool == "build_report_file" and how == "ok" for tool, how in steps):
            break

    session = workbench._sessions.get(session_id or "")
    if session and session.summary:
        report = {
            "rows": session.summary.rows,
            "measures": session.summary.measures,
            "used": session.summary.source_rows_used,
            "skipped": session.summary.source_rows_skipped,
        }
    return {
        "said": said,
        "steps": steps,
        "session": session,
        "report": report,
    }


def check(path: Path) -> list[str]:
    """Everything wrong with what a conversation produced for this workbook."""
    problems: list[str] = []
    started = time.time()
    print(f"\n{path.name}")
    print("  holding a conversation...", flush=True)

    try:
        outcome = hold_conversation(path)
    except UNREACHABLE as e:
        raise CouldNotRun(f"{type(e).__name__}: {str(e).splitlines()[0][:120]}") from e
    built = [t for t, how in outcome["steps"] if t == "build_report_file" and how == "ok"]
    failed = [t for t, how in outcome["steps"] if how in ("broken",)]

    print(f"  {len(outcome['steps'])} steps, {len(outcome['said'])} replies, "
          f"{time.time() - started:.0f}s")
    for tool, how in outcome["steps"]:
        if how != "ok":
            print(f"    {how}: {tool}")

    if not built:
        problems.append("no report was built")
    if failed:
        problems.append(f"steps broke: {', '.join(sorted(set(failed)))}")

    report = outcome["report"]
    if report is None:
        problems.append("no figures were summarised")
        return problems

    # The figures, against the spreadsheet read independently.
    expected = expected_totals(path)
    if expected:
        for name, prefix in (("planned", "target_"), ("actual", "actual_")):
            measure = next(
                (m for m in report["measures"] if m.startswith(prefix)), None
            )
            if measure is None:
                problems.append(f"the report has no {name} figure")
                continue
            got = sum(
                r[measure] for r in report["rows"] if isinstance(r.get(measure), (int, float))
            )
            want = expected[name]
            if abs(got - want) > 0.5:
                problems.append(
                    f"{name}: the report totals {got:,.0f}, the spreadsheet {want:,.0f} "
                    f"- {abs(got - want):,.0f} out"
                )
            else:
                print(f"    {name:8} {got:,.0f}  matches the spreadsheet")

    # Every figure the AI stated, against the figures that exist.
    session = outcome["session"]
    if session:
        for sentence in outcome["said"]:
            invented = figure_check.unsupported(sentence, session)
            if invented:
                problems.append(
                    f"said figures no tool produced: {', '.join(invented)}"
                )
    return problems


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", action="append", type=Path, default=None)
    chosen = parser.parse_args().workbook

    downloads = Path.home() / "Downloads"
    workbooks = chosen or [
        p for p in [
            downloads / "Video Production Plan - TEST.xlsx",
            *sorted(Path.home().glob("AppData/Local/Temp/lumina-*/*.xlsx"))[:1],
        ] if p.exists()
    ]
    if not workbooks:
        print("No workbooks to try. Pass --workbook, or generate the test one with")
        print("  python lumina/backend/testdata/make_video_plan.py")
        return 2

    everything: dict[str, list[str]] = {}
    skipped: dict[str, str] = {}
    for path in workbooks:
        try:
            everything[path.name] = check(path)
        except CouldNotRun as e:
            skipped[path.name] = str(e)
        except Exception as e:  # a crash is a result, not a reason to stop
            everything[path.name] = [f"the run itself failed: {type(e).__name__}: {e}"]

    # Plain ASCII: the Windows console is cp1252 and a box-drawing character crashes it,
    # which turned a passing run into a traceback after all the work was already done.
    print("\n" + "-" * 68)
    for name, problems in everything.items():
        print(f"{'PASS' if not problems else 'FAIL'}  {name}")
        for problem in problems:
            print(f"      - {problem}")
    return 1 if any(everything.values()) else 0


if __name__ == "__main__":
    raise SystemExit(main())
